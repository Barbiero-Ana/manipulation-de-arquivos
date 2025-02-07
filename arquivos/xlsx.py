'''
6. Criar uma planilha com dados fictícios
● Crie um programa que:
○ Gere um arquivo Excel chamado dados.xlsx.
○ Crie uma planilha chamada "Vendas".
○ Adicione cabeçalhos nas colunas (ex.: "Produto", "Quantidade", "Preço").
○ Adicione 5 linhas de dados fictícios.
○ Salve o arquivo.


'''

from openpyxl import Workbook
from openpyxl import load_workbook #lê o arquivo xlsx no terminal de python


wb = Workbook()
ws = wb.active
ws.title = 'Vendas'
ws.append(['Produto', 'Quantidade', 'Preço'])

while True:
    print('Deseja adicionar os dados à planilha de vendas?\n1 - Sim\n2 - Não')
    decisao = int(input('- '))
    
    if decisao == 1:
        for i in range(5):
            print(f'Insira os dados para a linha {i+1}: ')
            produto = input('Nome do produto - ')
            quantidade = int(input('Quantidade do produto - '))
            preco = float(input('Preço do produto - '))
            ws.append([produto, quantidade, preco])  
        
        wb.save('vendas.xlsx')  
        print('Planilha "vendas.xlsx" criada com sucesso!')
        break

    elif decisao == 2:
        print('Ok, continuando o sistema...')
        break

    else:
        print('Opção inválida...')


# 7 - abra o arquivo xlsx, leia o todo e imprima seus dados no terminal

wb = load_workbook('vendas.xlsx')
ws = wb.active
print('\nDados contidos na planilha:')
for row in ws.iter_rows(values_only=True):
    print(row)


# 8 - Filtrar dados dentro da panilha, leia a planilha e os dados de vendas, filtre as linhas no qual o valor do produto seja mais que R$ 100,00 - Salve essas linhas (ou essa linha n sei) em uma nova planilha chamada de vendas_filtradas.xlsx

while True:
    print('\nDeseja filtrar vendas com preço maior que R$100,00?\n1 - Sim\n2 - Não')
    decisao = int(input('- '))

    if decisao == 1:
        wb_new = Workbook()  
        ws_new = wb_new.active
        ws_new.title = 'Vendas Filtradas'

        ws_new.append(['Produto', 'Quantidade', 'Preço'])

        for row in ws.iter_rows(min_row=2, values_only=True):  
            produto, quantidade, preco = row  
            if preco > 100:
                ws_new.append([produto, quantidade, preco])

        wb_new.save('vendas_filtradas.xlsx')  
        print('Arquivo "vendas_filtradas.xlsx" criado com sucesso!')
        break

    elif decisao == 2:
        print('Ok, saindo do sistema e continuando o código...')
        break

    else:
        print('Opção inválida')